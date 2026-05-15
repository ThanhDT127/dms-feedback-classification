"""
Excel output formatter — merge major/minor headers with color coding.

Ported from notebook Cell 9 (Excel header format).
"""
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from pipeline.issue_classifier import MINOR_ORDER, MINOR_TO_MAJOR


# ── Color palette for major categories ───────────────────────────────────────
MAJOR_COLORS = {
    "Sản phẩm": "FFE699",
    "Yêu cầu công cụ BH": "C6E0B4",
    "Giá, cơ chế RD": "BDD7EE",
    "Dịch vụ": "F8CBAD",
    "Hàng giả": "F4B183",
    "Website": "D9E1F2",
    "Đối thủ cạnh tranh": "C9C9C9",
    "Tin trung lập": "FFD966",
}
BLUE_OTHER = "DDEBF7"


def write_formatted_header(ws, df_like: pd.DataFrame):
    """
    Write a two-row header to an openpyxl worksheet:
    - Row 1: Major category (merged across minor columns)
    - Row 2: Minor category labels

    Non-minor columns get a merged cell spanning both rows.

    Args:
        ws: openpyxl Worksheet object.
        df_like: DataFrame whose columns define the header structure.
    """
    header_font = Font(bold=True)
    thin = Side(style="thin", color="000000")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    ncols = len(df_like.columns)

    # ── Write header cells ──
    for col_idx, col_name in enumerate(df_like.columns, start=1):
        if col_name in MINOR_ORDER:
            # Minor column: row 1 = major, row 2 = minor
            ws.cell(row=1, column=col_idx, value=MINOR_TO_MAJOR[col_name]).font = header_font
            ws.cell(row=2, column=col_idx, value=col_name).font = header_font
            ws.cell(row=1, column=col_idx).alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            ws.cell(row=2, column=col_idx).alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
        else:
            # Non-minor: merge rows 1+2
            ws.merge_cells(start_row=1, end_row=2, start_column=col_idx, end_column=col_idx)
            c = ws.cell(row=1, column=col_idx, value=col_name)
            c.font = header_font
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── Merge major category blocks ──
    start_block, prev_major = None, None
    for col_idx, col_name in enumerate(df_like.columns, start=1):
        major = MINOR_TO_MAJOR.get(col_name) if col_name in MINOR_ORDER else None
        if major != prev_major:
            if prev_major and start_block:
                end_col = col_idx - 1
                ws.merge_cells(start_row=1, start_column=start_block, end_row=1, end_column=end_col)
            prev_major = major
            start_block = col_idx if major else None

    if prev_major and start_block:
        end_col = ncols
        while end_col >= start_block and df_like.columns[end_col - 1] not in MINOR_ORDER:
            end_col -= 1
        if end_col >= start_block:
            ws.merge_cells(start_row=1, start_column=start_block, end_row=1, end_column=end_col)

    # ── Apply colors ──
    for col_idx, col_name in enumerate(df_like.columns, start=1):
        if col_name in MINOR_ORDER:
            color = MAJOR_COLORS.get(MINOR_TO_MAJOR[col_name], "FFFFFF")
        else:
            color = BLUE_OTHER
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws.cell(row=1, column=col_idx).fill = fill
        ws.cell(row=2, column=col_idx).fill = fill
        ws.cell(row=1, column=col_idx).border = border_thin
        ws.cell(row=2, column=col_idx).border = border_thin
