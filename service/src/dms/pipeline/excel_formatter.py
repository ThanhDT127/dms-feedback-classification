"""Excel header formatting helpers."""

from __future__ import annotations

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .issue_classifier import MINOR_ORDER, MINOR_TO_MAJOR

MAJOR_COLORS = {
    "Sản phẩm": "FFE699",
    "Yêu cầu công cụ BH": "C6E0B4",
    "Giá, cơ chế RD": "BDD7EE",
    "Dịch vụ": "F8CBAD",
    "Hàng giả": "F4B183",
    "Website": "D9E1F2",
    "Đối thủ cạnh tranh": "C9C9C9",
    "Tin trung lập": "FFD966",
}
BLUE_OTHER = "DDEBF7"


def write_formatted_header(ws, df_like: pd.DataFrame) -> None:
    """Write a two-row grouped header into an openpyxl worksheet."""
    header_font = Font(bold=True)
    thin = Side(style="thin", color="000000")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    ncols = len(df_like.columns)
    for col_idx, col_name in enumerate(df_like.columns, start=1):
        if col_name in MINOR_ORDER:
            ws.cell(row=1, column=col_idx, value=MINOR_TO_MAJOR[col_name]).font = header_font
            ws.cell(row=2, column=col_idx, value=col_name).font = header_font
            ws.cell(row=1, column=col_idx).alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            ws.cell(row=2, column=col_idx).alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
        else:
            ws.merge_cells(start_row=1, end_row=2, start_column=col_idx, end_column=col_idx)
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    start_block, prev_major = None, None
    for col_idx, col_name in enumerate(df_like.columns, start=1):
        major = MINOR_TO_MAJOR.get(col_name) if col_name in MINOR_ORDER else None
        if major != prev_major:
            if prev_major and start_block:
                ws.merge_cells(start_row=1, start_column=start_block, end_row=1, end_column=col_idx - 1)
            prev_major = major
            start_block = col_idx if major else None

    if prev_major and start_block:
        end_col = ncols
        while end_col >= start_block and df_like.columns[end_col - 1] not in MINOR_ORDER:
            end_col -= 1
        if end_col >= start_block:
            ws.merge_cells(start_row=1, start_column=start_block, end_row=1, end_column=end_col)

    for col_idx, col_name in enumerate(df_like.columns, start=1):
        color = MAJOR_COLORS.get(MINOR_TO_MAJOR[col_name], "FFFFFF") if col_name in MINOR_ORDER else BLUE_OTHER
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws.cell(row=1, column=col_idx).fill = fill
        ws.cell(row=2, column=col_idx).fill = fill
        ws.cell(row=1, column=col_idx).border = border_thin
        ws.cell(row=2, column=col_idx).border = border_thin
